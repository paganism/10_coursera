import requests
from lxml import etree
from bs4 import BeautifulSoup
from openpyxl import Workbook
import argparse
import os


def get_courses_link_list(xml_content, record_count):
    root = etree.fromstring(xml_content)
    url_list = []
    for element in root.getchildren():
        for child in element.getchildren():
            url_list.append(child.text)
    return url_list[:record_count]


def get_content(link):
    response = requests.get(link).content
    return response


def get_course_inform(html_content, course):
    soup = BeautifulSoup(html_content, 'html.parser')
    course_inform = dict(
        course_title=soup.find_all('h2')[0].get_text(),
        language=soup.find_all(
            'div',
            'rc-Language'
        )[0].get_text(),
        start_date=soup.find_all(
            'div',
            'rc-StartDateString'
        )[0].get_text(),
        continuation=soup.find_all(
            'div',
            'rc-BasicInfo'
        )[0].get_text(),
        rating=None
    )
    rating = soup.find_all(
        'div',
        'ratings-text headline-2-text'
    )
    if rating:
        course_inform['rating'] = rating[0].getText()
    return course_inform


def output_courses_info_to_xls(course_list, ws):
    ws.title = 'Coursera courses info'
    ws['A1'] = 'course_title'
    ws['B1'] = 'language'
    ws['C1'] = 'start_date'
    ws['D1'] = 'continouation'
    ws['E1'] = 'rating'
    for course in course_list:
        ws.append([
            course['course_title'],
            course['language'],
            course['start_date'],
            course['continuation'],
            course['rating']
        ])
    return ws


def save_work_book(wb, filepath):
    wb.save(filepath)


def parse_argument():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '--path',
        dest='path',
        required=True,
        help='Path to file'
    )
    return parser.parse_args()


if __name__ == '__main__':
    record_count = 20
    link = 'https://www.coursera.org/sitemap~www~courses.xml'
    xml_content = get_content(link)
    url_list = get_courses_link_list(xml_content, record_count)
    course_list = []
    arg = parse_argument()
    filepath = arg.path
    for url in url_list:
        html_content = get_content(url)
        course_info = get_course_inform(html_content, url)
        course_list.append(course_info)
    wb = Workbook()
    ws = wb.active
    output_courses_info_to_xls(course_list, ws)
    save_work_book(wb, filepath)
